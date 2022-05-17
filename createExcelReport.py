from openpyxl import Workbook, load_workbook
from openpyxl.styles import PatternFill

wb = Workbook()


def create_excel_file():
    # especificamos el nombre y la ruta del archivo
    filesheet = "./san_and_storage_inventory.xlsx"
    # guardamos el archivo
    wb.save(filesheet)


def define_worksheets():
    filesheet = "./san_and_storage_inventory.xlsx"
    wb = load_workbook(filesheet)
    sheet1 = wb.create_sheet("ResumenStorageXespacio", 0)
    sheet2 = wb.create_sheet("All_San_Inventory", 1)
    wb.save(filesheet)


def fill_csv_in_list():
    my_list_of_csv = []
    csv_file = open("all_san_sw_inventory_2022-05-06.csv")

    for i in csv_file:
        my_list_of_csv.append(i.split(','))

    return my_list_of_csv


def fill_data_in_inventory(sheet_name, my_list_of_csv):
    filesheet = "./san_and_storage_inventory.xlsx"
    wb = load_workbook(filesheet)
    sheet = wb[sheet_name]

    for data in my_list_of_csv:
        sheet.append(data)

    wb.save(filesheet)


def get_cell_in_map(port_coord):
    sansw_directorMap = open("DirectorSheetMap.conf")
    for lines in sansw_directorMap:
        if port_coord in lines:
            lista = lines.split(';')
            return "{}".format(lista[1]).split('/')


def fill_sandir_inventory(sheet_name, my_list_of_csv, sansw_name1, sansw_name2):
    filesheet = "./san_and_storage_inventory.xlsx"
    wb = load_workbook(filesheet)
    sheet = wb[sheet_name]
    greenFill = PatternFill(start_color='79f009',
                   end_color='79f009',
                   fill_type='solid')

    for data in my_list_of_csv:
        #print(data)
        if data[1] == sansw_name1 and data[9] == 'Online' or data[1] == sansw_name2 and data[9] == 'Online':
            port_coord = "{}/{}".format(data[4], data[5])
            list = get_cell_in_map(port_coord)
            sheet['{}{}'.format(list[0], list[1].rstrip())].fill = greenFill

    wb.save(filesheet)


if __name__ == '__main__':
    #create_excel_file()
    #define_worksheets()
    #fill_data_in_inventory("All_San_Inventory", fill_csv_in_list())
    fill_sandir_inventory('SANSW1', fill_csv_in_list(), 'virtualSW1', 'virtualSW2')
    fill_sandir_inventory('SANSW2', fill_csv_in_list(), 'virtualSW1', 'virtualSW2')
